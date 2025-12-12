# Exportación de los datos a Excel utilizando "openpyxl" y "MySQL Connector/MySQL Python".
    # Módulo para exportar datos de la base de datos MySQL a un archivo Excel.
import os                                                               # Importa la biblioteca "os" para operaciones del sistema operativo.

from openpyxl import Workbook                                           # Importa la clase Workbook para crear archivos Excel.
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # Importa estilos para formato.
from db_connection import connection                                    # Importa la función de conexión a la base de datos desde el módulo "db_connection".




# Función/Método para exportar los datos a un archivo Excel.
def export_data_to_excel(host, database, user, password, cod_medico_inicio = None, cod_medico_fin = None):
    # Crea una carpeta/Directorio "Informes Excel generados" en el directorio actual.
    ruta_carpeta = os.path.join(os.getcwd(), "Informes Excel generados")    # Define la ruta de la carpeta.

    # Compreuba/Verifica previo a crear la carpeta si esta existe. Si no existe crea la carpeta.
    if not os.path.exists(ruta_carpeta):
        os.makedirs(ruta_carpeta)                       # Crea la carpeta si no existe.
        print(f"\n\n\tCarpeta creada: {ruta_carpeta}")  # Mensaje de confirmación de creación de la carpeta por consola.
    # Si la carpeta ya existe, muestra un mensaje indicando/informando al usuario de su existencia.
    else:
        print(f"\n\n\tLa carpeta ya existe: {ruta_carpeta}")    # Mensaje de confirmación de que la carpeta ya existe por consola.
    
    # Establece la conexión a la base de datos.
    conn = connection(host, database, user, password)

    # Crea un cursor para ejecutar consultas SQL.
    curSQL = conn.cursor()

    # Ejecuta una consulta SQL para obtener los datos de ambas tabla "MEDICOS" y "CITAS".
    querySQL = ("""
                    SELECT m.COD_MEDICO, m.NOMBRE_COMPLETO, m.ESPECIALIDAD, m.TURNO, m.CONSULTAS_DISPONIBLES_LUNES, m.CONSULTAS_DISPONIBLES_MARTES, m.CONSULTAS_DISPONIBLES_MIERCOLES, m.CONSULTAS_DISPONIBLES_JUEVES, m.CONSULTAS_DISPONIBLES_VIERNES, m.ANOS_EXPERIENCIA, c.NUM_CITA, c.COD_MEDICO, c.FECHA_CITA, c.HORA_CITA, c.MODALIDAD, c.URGENTE, c.ESTADO
                    FROM MEDICOS m
                    LEFT JOIN CITAS c ON m.COD_MEDICO = c.COD_MEDICO
                """)

    # Añade condiciones "WHERE" según los parámetros.
    condiciones = []
    parametros = []

    # Filtrado por código de médico.
    if cod_medico_inicio is not None:
        # Si se proporciona un código de médico de inicio.
        if cod_medico_fin is not None:
            # Rango de médicos (Parámetros Médico-inicio - Médico-fin).
            condiciones.append("m.COD_MEDICO BETWEEN %s AND %s")    # Añade la condición al listado de condiciones.
            parametros.extend([cod_medico_inicio, cod_medico_fin])  # Añade los parámetros correspondientes al listado de parámetros.
        # Si no se proporciona un código de médico de fin.
        else:
            # Rango de médicos (Parámetros Médico-inicio - x).
                # "x" = Parámetro Médico-no-específico. Por tanto se imprimirá (Parámetros Médico-inicio - hasta el final de la lista).
            condiciones.append("m.COD_MEDICO >= %s")    # Añade la condición al listado de condiciones.
            parametros.append(cod_medico_inicio)        # Añade el parámetro correspondiente al listado de parámetros.

    # Si se han añadido condiciones (parámetros de filtración de búsqueda), las agrega a la consulta SQL. Construye la consulta final con las condiciones.
    if condiciones:
        querySQL += " WHERE " + " AND ".join(condiciones)   # Añade las condiciones a la consulta SQL.
    
    querySQL += " ORDER BY m.COD_MEDICO, c.NUM_CITA"   # Añade el ordenamiento a la consulta SQL. La sentencia final quedará de la siguiente forma:
                                                            # SELECT ...
                                                            # FROM ...
                                                            # LEFT JOIN ...
                                                            # WHERE ... (si hay condiciones)
                                                            # ORDER BY ...
                                                    
    # Ejecuta la consulta SQL añadiendo la sentencia, los parámetros y condiciones.
    curSQL.execute(querySQL, parametros)

    # Obtiene todos los datos resultantes de la consulta.
    datos = curSQL.fetchall()

    # Una vez terminada la conexión y su posterior consulta de ejecución cierra la conexión a la base de datos.
    conn.close()

    # Define el nombre del archivo según el tipo de exportación.
    if cod_medico_inicio is None:
        excel_fileName = "informe_medicos_citas_completo.xlsx"  # Nombre del archivo Excel completo.
        titulo = "Informe Completo de Médicos y Citas"          # Título del informe completo.
    elif cod_medico_fin is not None:
        excel_fileName = f"informe_medicos_{cod_medico_inicio}_a_{cod_medico_fin}.xlsx" # Nombre del archivo Excel por rango de médicos.
        titulo = f"Informe de Médicos {cod_medico_inicio} a {cod_medico_fin}"           # Título del informe por rango de médicos.
    else:
        excel_fileName = f"informe_medico_{cod_medico_inicio}.xlsx" # Nombre del archivo Excel por médico específico.
        titulo = f"Informe del Médico {cod_medico_inicio}"          # Título del informe por médico específico.
    
    # Define la ruta completa del archivo Excel.
    excel_file = os.path.join(ruta_carpeta, excel_fileName)

    # Crea el libro de trabajo (workbook) de Excel.
    wb = Workbook()                 # Crea un nuevo libro de trabajo.
    ws = wb.active                  # Selecciona la hoja activa.
    ws.title = "Informe Médicos"    # Título de la hoja de cálculo.

    # Define los estilos para el formato del Excel.
    header_font = Font(name = 'Arial', size = 12, bold = True, color = "FFFFFF")                    # Define el estilo de fuente para los encabezados.
    header_fill = PatternFill(start_color = "2C3E50", end_color = "2C3E50", fill_type = "solid")    # Define el estilo de relleno para los encabezados.
    header_alignment = Alignment(horizontal = "center", vertical = "center", wrap_text = True)      # Define la alineación para los encabezados.
    
    # Estilos para el título principal.
    title_font = Font(name = 'Arial', size = 16, bold = True, color = "2C3E50") # Define el estilo de fuente para el título principal.
    title_alignment = Alignment(horizontal = "center", vertical = "center")     # Define la alineación para el título principal.
    
    # Estilo de borde para las celdas.
    border_style = Border(
        left = Side(style = 'thin'),
        right = Side(style = 'thin'),
        top = Side(style = 'thin'),
        bottom = Side(style = 'thin')
    )

    # Agregar título principal.
    ws.merge_cells('A1:Q1')                 # Fusiona/Combina las celdas de la A1 a la Q1 para el título.
    ws['A1'] = titulo                       # Asigna el título a la celda A1.
    ws['A1'].font = title_font              # Aplica el estilo de fuente al título.
    ws['A1'].alignment = title_alignment    # Alinea el título en el centro.
    ws.row_dimensions[1].height = 30        # Ajusta la altura de la fila del título.

    # Define los encabezados.
    encabezados = [
        'Código Médico', 'Nombre Completo', 'Especialidad', 'Turno',
        'Consultas L', 'Consultas M', 'Consultas X', 'Consultas J', 'Consultas V',
        'Años Experiencia', 'Nº Cita', 'Cód. Médico Cita', 'Fecha Cita',
        'Hora Cita', 'Modalidad', 'Urgente', 'Estado'
    ]

    # Escribe encabezados en la fila 3.
    for col_num, encabezado in enumerate(encabezados, 1):   # Comienza desde la columna 1 (A).
        cell = ws.cell(row = 3, column = col_num)           # Selecciona la celda correspondiente.
        cell.value = encabezado                             # Asigna el valor del encabezado a la celda.
        cell.font = header_font                             # Aplica el estilo de fuente a los encabezados.
        cell.fill = header_fill                             # Aplica el estilo de relleno a los encabezados.
        cell.alignment = header_alignment                   # Alinea los encabezados en el centro.
        cell.border = border_style                          # Aplica el estilo de borde a los encabezados.

    ws.row_dimensions[3].height = 35                        # Ajusta la altura de la fila de los encabezados.

    # Verifica/Comprueba si la lista de datos está vacía. Si no hay datos, escribie el mensaje.
    if not datos:
        ws.merge_cells('A4:Q4')                                                             # Fusiona/Combina las celdas de la A4 a la Q4 para el mensaje.
        ws['A4'] = "No existen o no se encontraron datos para los criterios especificados." # Mensaje cuando no hay datos.
        ws['A4'].alignment = Alignment(horizontal = "center", vertical = "center")          # Alinea el mensaje en el centro.
        ws['A4'].font = Font(name = 'Arial', size = 11, italic = True)                      # Aplica el estilo de fuente al mensaje.
    # Si hay datos, escribie los datos en las filas correspondientes.
    else:
        # Escribie los datos a partir de la fila 4.
        row_num = 4

        # Recorre los datos y los escribe en las celdas correspondientes.
        for row_data in datos:
            # Recorre cada valor en la fila de datos.
            for col_num, value in enumerate(row_data, 1):
                # Selecciona la celda correspondiente.
                cell = ws.cell(row = row_num, column = col_num)
                
                # Verifica/Comprueba si no hay datos. Si no hay datos formatea los valores "None".
                if value is None:
                    cell.value = "Sin datos"
                # Si el valor es de tipo fecha, lo formatea adecuadamente.
                else:
                    cell.value = str(value)
                
                # Aplica el estilo de borde a las celdas. Alineando los datos a la izquierda y centrado verticalmente
                cell.border = border_style
                cell.alignment = Alignment(horizontal = "left", vertical = "center")
                
                # Resalta las filas alternadas.
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color = "ECF0F1", end_color = "ECF0F1", fill_type = "solid")  # Relleno para filas pares.
            
            # Incrementa el número de fila para la siguiente iteración.
            row_num += 1

    # Ajusta el ancho de columnas automáticamente.
    for column in ws.columns:
        # Calcula el ancho máximo necesario para cada columna.
        max_length = 0                          # Inicializa la longitud máxima.
        column_letter = column[0].column_letter # Obtiene la letra de la columna.

        # Recorre cada celda en la columna para encontrar la longitud máxima del contenido.
        for cell in column:
            try:
                # Verifica/Comprueba si la longitud del valor de la celda es mayor que la longitud máxima actual.
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass    # Ignora errores (por ejemplo, celdas vacías).

        # Ajusta el ancho de la columna con un límite máximo.
        adjusted_width = min(max_length + 2, 50)                    # Máximo 50 caracteres de ancho.
        ws.column_dimensions[column_letter].width = adjusted_width  # Ajusta el ancho de la columna.

    # Congela los paneles (fijar encabezados).
    ws.freeze_panes = 'A4'

    # Guarda el archivo Excel.
    wb.save(excel_file)

    # Mensaje de confirmación de la ejecución por consola.
    print(f"\n\n\tLos datos han sido exportados exitosamente a {excel_file}")
    
    # Devuelve el nombre del archivo Excel generado.
    return excel_file